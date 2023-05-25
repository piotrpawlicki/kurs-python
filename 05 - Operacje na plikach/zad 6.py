import openpyxl
import datetime
from tkinter import Tk
from tkinter.filedialog import askopenfilename


def read_from_xlsx():
    input('Press enter to open file with card numbers:  \n')
    Tk().withdraw()
    filename = askopenfilename()
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    return sheet


def list_of_cards(sheet):
    cards = []
    for row in sheet.iter_rows():
        cards.append(row[0].value)
    return cards


def write_list_to_column(sheet, column_index, column_name, cards):
    column = column_index + '1' # okreslenie komorki startowej dla kolumny\
    sheet[column] = column_name #dodanie nazwy kolumny
    for i, card in enumerate(cards, start = 2): #zaczynamy od elementu 2, bo pierwszy jest zajęty dla nazwy kolumny a 0 w excelu nie istnieje
        cell = column_index + str(i)
        sheet[cell] = card


def visa(number):
    return str(number).startswith('4') and len(number) in (13, 16)


def amex(number):
    return str(number).startswith('3') and len(number) == 15 and number[1] in ('4', '7')


def master_card(number):
    return len(number) == 16 and (
        int(number[0:2]) in range(51,56) or
        2221 <= int(number[:4]) <= 2720
    )


def check_number(num_list):
    visa_list=[]
    amex_list=[]
    master_card_list=[]
    for i in range(len(num_list)):
        num = str(num_list[i])
        if visa(num):
            visa_list.append(num)
        elif amex(num):
             amex_list.append(num)
        elif master_card(num):
            master_card_list.append(num)
    return visa_list, amex_list, master_card_list


def main():
    sheet = read_from_xlsx()
    cards = list_of_cards(sheet)
    visa_list, amex_list, master_card_list = check_number(cards)
    input('Press enter to open file where I should write to:  \n')
    filename = askopenfilename()
    wb = openpyxl.load_workbook(filename)
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  ##użycie formatowania do timestammp,. aby unikną podawania w nazwie worksheeta zabronionego ' : '
    sheet_name = f'result {timestamp}'
    sheet = wb.create_sheet(sheet_name)
    write_list_to_column(sheet, 'A', 'Visa', visa_list)
    write_list_to_column(sheet, 'B', 'Amex', amex_list)
    write_list_to_column(sheet, 'C', 'MasterCard', master_card_list)
    wb.save(filename)


main()